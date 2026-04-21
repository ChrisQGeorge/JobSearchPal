"""Excel template + import helpers for TrackedJob bulk operations.

The template is a single-sheet .xlsx with one header row and one sample row.
Column names match TrackedJob fields where possible. The importer is lenient:
unknown headers are ignored, blank cells → None, case-insensitive status /
enum normalization, CSV-splitting for skills / tech-stack-hints.
"""

from __future__ import annotations

import io
import logging
from datetime import date, datetime
from typing import Any, Iterable, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)


# Ordered columns in the template. Keep these stable — users may hand-edit
# templates they've downloaded earlier.
TEMPLATE_COLUMNS: list[tuple[str, str, str]] = [
    # (header, TrackedJob field key, example cell value)
    ("Title", "title", "Senior Widget Engineer"),
    ("Organization", "organization_name", "Acme Corp"),
    ("Status", "status", "applied"),
    ("Priority", "priority", "high"),
    ("Location", "location", "Remote · US"),
    ("Remote policy", "remote_policy", "remote"),
    ("Employment type", "employment_type", "full_time"),
    ("Experience level", "experience_level", "senior"),
    ("Years experience min", "experience_years_min", 5),
    ("Years experience max", "experience_years_max", 8),
    ("Education required", "education_required", "bachelors"),
    ("Visa sponsorship (yes/no)", "visa_sponsorship_offered", "yes"),
    ("Relocation offered (yes/no)", "relocation_offered", "no"),
    ("Salary min", "salary_min", 150000),
    ("Salary max", "salary_max", 200000),
    ("Salary currency", "salary_currency", "USD"),
    ("Source URL", "source_url", "https://example.com/jobs/123"),
    ("Source platform", "source_platform", "linkedin"),
    ("Date posted", "date_posted", date(2026, 4, 1)),
    ("Date discovered", "date_discovered", date(2026, 4, 10)),
    ("Date applied", "date_applied", date(2026, 4, 12)),
    ("Date closed", "date_closed", ""),
    ("Required skills (comma-separated)", "required_skills", "Python, FastAPI, PostgreSQL"),
    ("Nice-to-have skills (comma-separated)", "nice_to_have_skills", "Kubernetes, GraphQL"),
    ("Notes", "notes", "Got intro from Jamie last Tuesday."),
    ("Job description", "job_description", "Paste the full JD here."),
]


def build_template_workbook() -> bytes:
    """Return a fresh template .xlsx as bytes — headers + one sample row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    header_fill = PatternFill("solid", fgColor="1F2430")
    header_font = Font(bold=True, color="E8C468")
    sample_font = Font(italic=True, color="8A94A6")

    for col_idx, (header, _field, sample) in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

        sample_cell = ws.cell(row=2, column=col_idx, value=sample)
        sample_cell.font = sample_font

        # Approximate width based on sample + header length.
        width = max(len(str(header)), len(str(sample)) if sample else 0) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width, 12), 50)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Second sheet with instructions + the enum vocabulary.
    notes = wb.create_sheet("Instructions")
    notes["A1"] = "Job Search Pal — Bulk Import Template"
    notes["A1"].font = Font(bold=True, size=14, color="E8C468")
    instructions = [
        "",
        "1. The first row is a sample — delete it before uploading (or leave it; the importer treats its values as literal).",
        "2. Status must be one of: watching, interested, applied, responded, screening, interviewing, assessment, offer, won, lost, withdrawn, ghosted, archived.",
        "3. Priority: low / medium / high.",
        "4. Remote policy: onsite / hybrid / remote.",
        "5. Experience level: junior / mid / senior / staff / principal / manager / director / vp / cxo.",
        "6. Employment type: full_time / part_time / contract / c2h / internship / freelance.",
        "7. Education required: none / associates / bachelors / masters / phd.",
        "8. Visa sponsorship + Relocation: yes / no / blank.",
        "9. Dates: YYYY-MM-DD or leave blank.",
        "10. Organization: free text. The importer resolves or creates it by name.",
        "11. Skills: comma-separated values in a single cell.",
        "12. Unknown columns are ignored, so feel free to add your own notes columns.",
    ]
    for i, line in enumerate(instructions, start=2):
        notes[f"A{i}"] = line
        notes[f"A{i}"].alignment = Alignment(wrap_text=True, vertical="top")
    notes.column_dimensions["A"].width = 110

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Import ----------------------------------------------------------------


_BOOL_TRUE = {"yes", "y", "true", "1", "offered", "sponsored"}
_BOOL_FALSE = {"no", "n", "false", "0", "not offered", "denied"}


def _parse_bool(v: Any) -> Optional[bool]:
    if v is None or v == "":
        return None
    s = str(v).strip().lower()
    if s in _BOOL_TRUE:
        return True
    if s in _BOOL_FALSE:
        return False
    return None


def _parse_date(v: Any) -> Optional[date]:
    if v in (None, ""):
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_number(v: Any) -> Optional[float]:
    if v in (None, ""):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("$", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_int(v: Any) -> Optional[int]:
    f = _parse_number(v)
    return int(f) if f is not None else None


def _parse_csv(v: Any) -> Optional[list[str]]:
    if v in (None, ""):
        return None
    return [s.strip() for s in str(v).split(",") if s.strip()]


def _cell(v: Any) -> Any:
    """Normalize empty strings and whitespace to None."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return v


def parse_workbook(data: bytes) -> list[dict]:
    """Parse uploaded .xlsx bytes into a list of field dicts suitable for
    creating TrackedJob records. Unknown columns are ignored; blank rows are
    skipped. Returns one dict per non-empty row.
    """
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb["Jobs"] if "Jobs" in wb.sheetnames else wb.active

    # Build header → field_key map. Case- and whitespace-insensitive.
    header_to_field: dict[int, str] = {}
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    header_lookup = {h.lower(): f for h, f, _ in TEMPLATE_COLUMNS}
    for col_idx, header in enumerate(first_row, start=1):
        if not header:
            continue
        key = header_lookup.get(str(header).strip().lower())
        if key:
            header_to_field[col_idx] = key

    if not header_to_field:
        raise ValueError(
            "No recognized header columns. Download the template and check the header row."
        )

    out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record: dict[str, Any] = {}
        for col_idx, value in enumerate(row, start=1):
            field = header_to_field.get(col_idx)
            if not field:
                continue
            record[field] = _cell(value)
        if not any(v is not None for v in record.values()):
            continue

        # Per-field parsing / coercion.
        if "title" in record and record["title"]:
            record["title"] = str(record["title"]).strip()
        for fk in ("experience_years_min", "experience_years_max"):
            if fk in record:
                record[fk] = _parse_int(record[fk])
        for fk in ("salary_min", "salary_max"):
            if fk in record:
                record[fk] = _parse_number(record[fk])
        for fk in (
            "date_posted",
            "date_discovered",
            "date_applied",
            "date_closed",
        ):
            if fk in record:
                record[fk] = _parse_date(record[fk])
        for fk in ("visa_sponsorship_offered", "relocation_offered"):
            if fk in record:
                record[fk] = _parse_bool(record[fk])
        for fk in ("required_skills", "nice_to_have_skills"):
            if fk in record:
                record[fk] = _parse_csv(record[fk])
        # Normalize enum fields to lowercase where present.
        for fk in (
            "status",
            "priority",
            "remote_policy",
            "employment_type",
            "experience_level",
            "education_required",
            "source_platform",
            "salary_currency",
        ):
            if record.get(fk) and isinstance(record[fk], str):
                record[fk] = record[fk].strip().lower()
        if record.get("salary_currency"):
            record["salary_currency"] = record["salary_currency"].upper()

        out.append(record)
    return out


# ============================================================================
# Queue-only import: a minimal template that just takes a job URL plus
# optional date fields and feeds the fetch queue instead of creating jobs
# directly. The queue worker then runs the URL fetcher to populate everything
# else.
# ============================================================================


QUEUE_TEMPLATE_COLUMNS: list[tuple[str, str, Any]] = [
    # (header, JobFetchQueue field key, example cell)
    ("Job URL", "url", "https://example.com/jobs/123"),
    ("Applied date (optional)", "desired_date_applied", date(2026, 4, 12)),
    ("Posted date (optional)", "desired_date_posted", date(2026, 4, 1)),
]


def build_queue_template_workbook() -> bytes:
    """Minimal queue-import template — URL + two optional dates."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Queue"

    header_fill = PatternFill("solid", fgColor="1F2430")
    header_font = Font(bold=True, color="E8C468")
    sample_font = Font(italic=True, color="8A94A6")

    for col_idx, (header, _field, sample) in enumerate(QUEUE_TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

        sample_cell = ws.cell(row=2, column=col_idx, value=sample)
        sample_cell.font = sample_font

        width = max(len(str(header)), len(str(sample)) if sample else 0) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width, 16), 60)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    notes = wb.create_sheet("Instructions")
    notes["A1"] = "Job Search Pal — Queue Import Template"
    notes["A1"].font = Font(bold=True, size=14, color="E8C468")
    instructions = [
        "",
        "1. Each non-empty row is added to the Fetch Queue — the Companion will visit the URL, extract job details, and create a TrackedJob in the background.",
        "2. Only the Job URL is required. Dates are optional.",
        "3. Applied date: if present, the created TrackedJob's date_applied is set to this value and its status is left at the default (watching) unless the Companion chooses otherwise from the page. To also mark the job as applied, use the single-URL queue form on the tracker page.",
        "4. Posted date: if present, this overrides any date the fetcher would extract from the posting.",
        "5. Dates accept YYYY-MM-DD, MM/DD/YYYY, or Excel's native date cells.",
        "6. The first row under the header is a sample — delete it before uploading, or leave it to import literally.",
        "7. Rows with an unrecognizable URL (not http:// or https://) are reported back as errors.",
    ]
    for i, line in enumerate(instructions, start=2):
        notes[f"A{i}"] = line
        notes[f"A{i}"].alignment = Alignment(wrap_text=True, vertical="top")
    notes.column_dimensions["A"].width = 110

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_queue_workbook(data: bytes) -> list[dict]:
    """Parse the queue-import workbook. Returns one dict per non-empty row.

    Each dict has at least a `url` key; date keys are included only when
    present. Unknown columns are ignored; blank rows skipped.
    """
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb["Queue"] if "Queue" in wb.sheetnames else wb.active

    header_lookup = {h.lower(): f for h, f, _ in QUEUE_TEMPLATE_COLUMNS}
    header_to_field: dict[int, str] = {}
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    for col_idx, header in enumerate(first_row, start=1):
        if not header:
            continue
        key = header_lookup.get(str(header).strip().lower())
        if key:
            header_to_field[col_idx] = key

    if not header_to_field:
        raise ValueError(
            "No recognized header columns. Download the queue template and "
            "check the header row."
        )

    out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record: dict[str, Any] = {}
        for col_idx, value in enumerate(row, start=1):
            field = header_to_field.get(col_idx)
            if not field:
                continue
            record[field] = _cell(value)
        if not any(v is not None for v in record.values()):
            continue
        if record.get("url"):
            record["url"] = str(record["url"]).strip()
        for fk in ("desired_date_applied", "desired_date_posted"):
            if fk in record:
                record[fk] = _parse_date(record[fk])
        out.append(record)
    return out
